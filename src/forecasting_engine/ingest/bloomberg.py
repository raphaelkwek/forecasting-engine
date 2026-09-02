"""Turning Bloomberg terminal exports into one contract-shaped CSV.

A Bloomberg export is one workbook per security: a ``Data`` sheet of dates and
fields, and a ``Metadata`` sheet naming what was pulled. The engine wants a
single CSV with one column per signal, so this module reads a pile of exports
and joins them on date.

**Files are identified by their ticker, never by their filename.** The first
real export we saw was named ``credit_spread_Data_LF98OAS_Index__values.xlsx``
and contained ``LF98TRUU`` — the high yield *total return* index rather than the
option-adjusted spread the contract asks for. The filename was wrong and the
metadata was right, so the metadata is what we trust.

That mistake is also why :func:`suspicious` exists. A total-return index in a
spread column is not a parse error and not a missing value; every cell reads
fine and every one of them is wrong. The only signal is that the numbers sit far
outside the documented range, so the converter says so loudly rather than
handing on a file that will fail validation for reasons nobody can trace back
to the export.

Gaps are left as gaps. Different indices keep different trading calendars, and
the union of their dates has holes in most columns. Forward-filling here would
hide from the quality report exactly what it exists to report.
"""

from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pandas as pd

from forecasting_engine.ingest.schema import COLUMNS, DATE_COLUMN, REQUIRED_COLUMNS

DATA_SHEET = "Data"
METADATA_SHEET = "Metadata"
DEFAULT_FIELD = "PX_LAST"

#: Bloomberg security → the contract column it supplies. Keyed on the ticker in
#: the Metadata sheet, because filenames have proven unreliable.
TICKER_MAP: Mapping[str, str] = {
    "SPX Index": "spx_close",
    "LEGATRUU Index": "agg_close",
    "VIX Index": "vix",
    "LF98OAS Index": "credit_spread_hy",
    "LUACOAS Index": "credit_spread_ig",
    "JPMVXYG7 Index": "fx_impl_vol",
    "USGGBE10 Index": "breakeven_10y",
}

#: Columns computed from other exports rather than pulled directly.
#: ``term_spread`` has no single Bloomberg ticker; it is the 10-year yield minus
#: the 2-year, so both legs must be exported and the difference taken here.
DERIVED: Mapping[str, tuple[str, str]] = {
    "term_spread": ("USGG10YR Index", "USGG2YR Index"),
}

#: Tickers close enough to a wanted one to be worth naming when they turn up.
NEAR_MISSES: Mapping[str, str] = {
    "LF98TRUU Index": (
        "the high yield total return index, not its spread — "
        "re-export LF98OAS Index for credit_spread_hy"
    ),
    "JPMVXYGL Index": (
        "the global FX volatility index, not the G7 one — "
        "re-export JPMVXYG7 Index for fx_impl_vol"
    ),
}


@dataclass(frozen=True)
class BloombergExport:
    """One workbook: what it holds, and where it came from."""

    path: Path
    security: str
    field: str
    series: pd.Series
    """Values indexed by date, ascending."""

    @property
    def column(self) -> str | None:
        """The contract column this supplies, or None if we do not want it."""
        return TICKER_MAP.get(self.security)


@dataclass(frozen=True)
class ConversionReport:
    """What the conversion did, and what a person still has to fix."""

    used: dict[str, str]
    """Contract column → the security that supplied it."""

    skipped: list[str]
    """One line per export we could not place, saying why."""

    missing: list[str]
    """Required columns nothing supplied."""

    warnings: list[str]
    """Values that parsed but look like the wrong Bloomberg field."""

    rows: int = 0

    @property
    def complete(self) -> bool:
        return not self.missing and not self.warnings

    def describe(self) -> str:
        lines = [f"Wrote {self.rows:,} rows covering {len(self.used)} of "
                 f"{len(REQUIRED_COLUMNS) - 1} signals."]
        for column, security in sorted(self.used.items()):
            lines.append(f"  ok       {column:<18} {security}")
        for note in self.skipped:
            lines.append(f"  skipped  {note}")
        for column in self.missing:
            lines.append(f"  MISSING  {column:<18} no export supplied this")
        for note in self.warnings:
            lines.append(f"  SUSPECT  {note}")
        return "\n".join(lines)


def read_export(path: Path, field: str = DEFAULT_FIELD) -> BloombergExport:
    """Read one Bloomberg workbook.

    Raises ``ValueError`` if the sheets or the field are not what we expect,
    naming what was there instead.
    """
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if DATA_SHEET not in book.sheetnames:
            raise ValueError(
                f"{path.name} has no {DATA_SHEET!r} sheet (found {book.sheetnames})"
            )
        rows = list(book[DATA_SHEET].iter_rows(values_only=True))
        security = _security(book)
    finally:
        book.close()

    if not rows:
        raise ValueError(f"{path.name} has an empty {DATA_SHEET!r} sheet")

    header = [str(cell) for cell in rows[0]]
    if field not in header:
        raise ValueError(f"{path.name} has no {field!r} column (found {header})")

    return BloombergExport(
        path=path,
        security=security,
        field=field,
        series=_series(rows, header.index(field)),
    )


def combine(exports: Sequence[BloombergExport]) -> tuple[pd.DataFrame, ConversionReport]:
    """Join exports into one contract-shaped frame, and report on the result.

    Every date any export saw appears in the output. A signal with no value on
    a date is left blank, which the contract treats as missing data rather than
    as an error.
    """
    placed: dict[str, pd.Series] = {}
    supplied: dict[str, str] = {}
    legs: dict[str, pd.Series] = {}
    skipped: list[str] = []

    for export in exports:
        if export.security in _derived_tickers():
            legs[export.security] = export.series
            continue
        column = export.column
        if column is None:
            skipped.append(f"{export.path.name}: {_why_skipped(export.security)}")
            continue
        if column in placed:
            skipped.append(
                f"{export.path.name}: {column} already supplied by {supplied[column]}"
            )
            continue
        placed[column] = export.series
        supplied[column] = export.security

    derived_notes = _derive(placed, supplied, legs)
    skipped.extend(derived_notes)

    frame = _assemble(placed)
    return frame, ConversionReport(
        used=supplied,
        skipped=skipped,
        missing=[c for c in REQUIRED_COLUMNS if c != DATE_COLUMN and c not in placed],
        warnings=suspicious(frame),
        rows=len(frame),
    )


def suspicious(frame: pd.DataFrame) -> list[str]:
    """Columns whose values sit almost entirely outside the documented range.

    A handful of breaches is a market event worth keeping. Nearly every value
    out of range means the wrong Bloomberg field was exported.
    """
    notes: list[str] = []
    for spec in COLUMNS:
        if spec.name not in frame.columns or spec.minimum is None and spec.maximum is None:
            continue
        values = pd.to_numeric(frame[spec.name], errors="coerce").dropna()
        if values.empty:
            continue
        outside = pd.Series(False, index=values.index)
        if spec.minimum is not None:
            outside |= values < spec.minimum
        if spec.maximum is not None:
            outside |= values > spec.maximum
        share = outside.mean()
        if share >= 0.9:
            notes.append(
                f"{spec.name}: {int(outside.sum()):,} of {len(values):,} values "
                f"({share:.0%}) fall outside the documented range "
                f"{_range(spec)} — observed {values.min():,.2f} to {values.max():,.2f}. "
                "This usually means the wrong Bloomberg field was exported."
            )
    return notes


def convert(
    paths: Iterable[Path], field: str = DEFAULT_FIELD
) -> tuple[pd.DataFrame, ConversionReport]:
    """Read every workbook in ``paths`` and combine them."""
    exports = []
    skipped: list[str] = []
    for path in paths:
        try:
            exports.append(read_export(path, field))
        except ValueError as exc:
            skipped.append(str(exc))
    frame, report = combine(exports)
    return frame, ConversionReport(
        used=report.used,
        skipped=skipped + report.skipped,
        missing=report.missing,
        warnings=report.warnings,
        rows=report.rows,
    )


def _security(book: openpyxl.Workbook) -> str:
    if METADATA_SHEET not in book.sheetnames:
        return ""
    for row in book[METADATA_SHEET].iter_rows(values_only=True):
        if row and str(row[0]).strip() == "Security":
            return str(row[1]).strip()
    return ""


def _series(rows: Sequence[tuple], position: int) -> pd.Series:
    dates: list[str] = []
    values: list[float] = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        parsed = pd.to_datetime(str(row[0])[:10], errors="coerce", format="ISO8601")
        if pd.isna(parsed):
            continue
        try:
            values.append(float(row[position]))
        except (TypeError, ValueError):
            continue  # #N/A and friends are missing data, not values
        dates.append(parsed.date().isoformat())
    return pd.Series(values, index=pd.Index(dates, name=DATE_COLUMN)).sort_index()


def _derived_tickers() -> frozenset[str]:
    return frozenset(t for legs in DERIVED.values() for t in legs)


def _derive(
    placed: dict[str, pd.Series], supplied: dict[str, str], legs: dict[str, pd.Series]
) -> list[str]:
    """Fill derived columns from their legs, reporting any that are incomplete."""
    notes: list[str] = []
    for column, (long_leg, short_leg) in DERIVED.items():
        have = [t for t in (long_leg, short_leg) if t in legs]
        if not have:
            continue
        if len(have) < 2:
            absent = long_leg if long_leg not in legs else short_leg
            notes.append(f"{column}: needs {long_leg} and {short_leg}; {absent} was not supplied")
            continue
        placed[column] = legs[long_leg] - legs[short_leg]
        supplied[column] = f"{long_leg} minus {short_leg}"
    return notes


def _assemble(placed: Mapping[str, pd.Series]) -> pd.DataFrame:
    """One row per date any export saw, in contract column order."""
    if not placed:
        return pd.DataFrame(columns=[DATE_COLUMN])
    frame = pd.DataFrame(placed).sort_index()
    frame.index.name = DATE_COLUMN
    ordered = [c for c in REQUIRED_COLUMNS if c != DATE_COLUMN and c in frame.columns]
    return frame[ordered].reset_index()


def _why_skipped(security: str) -> str:
    if not security:
        return "no Security in its Metadata sheet, so it could not be identified"
    if security in NEAR_MISSES:
        return f"{security} is {NEAR_MISSES[security]}"
    return f"{security} is not a signal the contract asks for"


def _range(spec) -> str:
    low = "any" if spec.minimum is None else f"{spec.minimum:g}"
    high = "any" if spec.maximum is None else f"{spec.maximum:g}"
    return f"{low} to {high}"
